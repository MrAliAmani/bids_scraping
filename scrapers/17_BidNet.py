import sys
import os
import io
import shutil

# Set stdout encoding to UTF-8 - only do this once at the start
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Add the parent directory to the Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.utils import safe_move, play_notification_sound

import time
import random
import logging
import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Set
from dataclasses import dataclass, asdict
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
from openpyxl.utils import get_column_letter
import re


# JavaScript helper for getting bid details
GET_BID_DETAILS_SCRIPT = """
function getBidType() {
    // Try first location
    const labelXPath1 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[3]/span';
    const valueXPath1 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[3]/div/p';
    
    // Try second location
    const labelXPath2 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[4]/span';
    const valueXPath2 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[4]/div/p';
    
    try {
        // Check first location
        let labelResult = document.evaluate(labelXPath1, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
        let labelElement = labelResult.singleNodeValue;
        
        if (labelElement && labelElement.textContent.trim().includes('Solicitation Type')) {
            const valueResult = document.evaluate(valueXPath1, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            const valueElement = valueResult.singleNodeValue;
            if (valueElement) {
                return valueElement.textContent.trim();
            }
        }
        
        // Check second location
        labelResult = document.evaluate(labelXPath2, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
        labelElement = labelResult.singleNodeValue;
        
        if (labelElement && labelElement.textContent.trim().includes('Solicitation Type')) {
            const valueResult = document.evaluate(valueXPath2, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            const valueElement = valueResult.singleNodeValue;
            if (valueElement) {
                return valueElement.textContent.trim();
            }
        }
    } catch (e) {
        console.log('Error in getBidType:', e);
    }
    
    return "Not found";
}

function getSolicitationNumber() {
    // Try first location
    const labelXPath1 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[4]/span';
    const valueXPath1 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[4]/div/p';
    
    // Try second location
    const labelXPath2 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[5]/span';
    const valueXPath2 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[5]/div/p';
    
    try {
        // Check first location
        let labelResult = document.evaluate(labelXPath1, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
        let labelElement = labelResult.singleNodeValue;
        
        if (labelElement && labelElement.textContent.trim().includes('Solicitation Number')) {
            const valueResult = document.evaluate(valueXPath1, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            const valueElement = valueResult.singleNodeValue;
            if (valueElement) {
                return valueElement.textContent.trim();
            }
        }
        
        // Check second location
        labelResult = document.evaluate(labelXPath2, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
        labelElement = labelResult.singleNodeValue;
        
        if (labelElement && labelElement.textContent.trim().includes('Solicitation Number')) {
            const valueResult = document.evaluate(valueXPath2, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            const valueElement = valueResult.singleNodeValue;
            if (valueElement) {
                return valueElement.textContent.trim();
            }
        }
    } catch (e) {
        console.log('Error in getSolicitationNumber:', e);
    }
    
    return "Not found";
}



function getBidDetails() {
    function trySelectors(selectors) {
        for (let selector of selectors) {
            try {
                // For CSS selector
                if (!selector.startsWith('/')) {
                    const element = document.querySelector(selector);
                    if (element) return element.textContent.trim();
                }
                // For XPath
                else {
                    const result = document.evaluate(
                        selector,
                        document,
                        null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE,
                        null
                    );
                    if (result.singleNodeValue) {
                        return result.singleNodeValue.textContent.trim();
                    }
                }
            } catch (e) {}
        }
        return "";
    }

    function getOrganizationName() {
        const selectors = {
            css: '#g_548 > div > p > a',
            xpath: '//*[@id="g_548"]/div/p/a',
            fullXpath: '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[2]/div/p/a'
        };

        for (let selectorType in selectors) {
            try {
                let element;
                const selector = selectors[selectorType];
                
                if (selectorType === 'css') {
                    element = document.querySelector(selector);
                } else {
                    const result = document.evaluate(
                        selector,
                        document,
                        null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE,
                        null
                    );
                    element = result.singleNodeValue;
                }

                if (element) {
                    return element.textContent.trim();
                }
            } catch (e) {}
        }
        return "";
    }

    function getContractingOfficeAddress() {
        const xpath1 = '//*[@id="g_632"]/div/p';
        const xpath2 = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[2]/div[1]/div/p';
        
        let element = document.evaluate(xpath1, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        if (element) return element.textContent.trim();
        
        element = document.evaluate(xpath2, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        if (element) return element.textContent.trim();
        
        return "";
    }

    function getContactInformation() {
        function getContentBySelectors(selectorInfo) {
            const selectors = {
                css: selectorInfo.css,
                xpath: selectorInfo.xpath,
                fullXpath: selectorInfo.fullXpath
            };

            for (let selectorType in selectors) {
                try {
                    let element;
                    const selector = selectors[selectorType];
                    
                    if (selectorType === 'css') {
                        element = document.querySelector(selector);
                    } else {
                        const result = document.evaluate(
                            selector,
                            document,
                            null,
                            XPathResult.FIRST_ORDERED_NODE_TYPE,
                            null
                        );
                        element = result.singleNodeValue;
                    }

                    if (element) {
                        return element.textContent.trim();
                    }
                } catch (e) {}
            }
            return "";
        }

        const name = getContentBySelectors({
            css: '#g_639 > div > p',
            xpath: '//*[@id="g_639"]/div/p',
            fullXpath: '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[5]/div[1]/div/p'
        });

        const phone = getContentBySelectors({
            css: '#g_640 > div > p',
            xpath: '//*[@id="g_640"]/div/p',
            fullXpath: '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[5]/div[2]/div/p'
        });

        const email = getContentBySelectors({
            css: '#g_641 > div > p',
            xpath: '//*[@id="g_641"]/div/p',
            fullXpath: '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[5]/div[3]/div/p'
        });

        return {
            name: name,
            title: "Contracting Officer",
            phone: phone,
            email: email
        };
    }

    function getTenderInstructions() {
        const selectors = {
            id: 'tenderInstructionsEN_0',
            xpath: '//*[@id="tenderInstructionsEN_0"]',
            fullXpath: '/html/body/main/div[1]/div[2]/div[3]/form/div[4]/div[2]/div/div/span'
        };

        for (let selectorType in selectors) {
            try {
                let element;
                const selector = selectors[selectorType];
                
                if (selectorType === 'id') {
                    element = document.getElementById(selector);
                } else {
                    const result = document.evaluate(
                        selector,
                        document,
                        null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE,
                        null
                    );
                    element = result.singleNodeValue;
                }

                if (element) {
                    // Try to click "See more" link if it exists
                    try {
                        const readMoreLink = element.querySelector('.read-more');
                        if (readMoreLink) {
                            readMoreLink.click();
                        }
                    } catch (e) {}

                    return element.textContent.replace('See more', '').trim();
                }
            } catch (e) {}
        }
        return "";
    }

    const details = {
        solicitation_number: getSolicitationNumber(),  // Use getSolicitationNumber() function
        title: trySelectors(['/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[5]/div/p']),
        agency: getOrganizationName(),
        notice_type: getBidType(),  // Use getBidType() function
        posted_date: trySelectors(['span.sol-publication-date span.date-value']),
        response_date: trySelectors(['span.sol-closing-date span.date-value']),
        description: trySelectors(['#descriptionText', '//*[@id="descriptionText"]']),
        contracting_office_address: getContractingOfficeAddress(),
        contact_info: getContactInformation(),
        additional_summary: getTenderInstructions()
    };

    return details;
}
return getBidDetails();
"""

GET_PUBLICATION_DATES_SCRIPT = """
function getAllPublicationDates() {
    const publicationDates = [];
    let rowIndex = 1;
    
    while (true) {
        const xpath = `/html/body/main/div[1]/div[1]/div/div/div/div[1]/div[2]/div/div/form/table/tbody/tr[${rowIndex}]/td/div/span/span/span[1]/span[2]`;
        try {
            const result = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            if (result.singleNodeValue) {
                const date = result.singleNodeValue.textContent.trim();
                publicationDates.push({
                    rowNumber: rowIndex,
                    publicationDate: date
                });
                rowIndex++;
            } else {
                break;
            }
        } catch (e) {
            break;
        }
    }
    return publicationDates;
}
return getAllPublicationDates();
"""

GET_CLOSING_DATES_SCRIPT = """
function getAllClosingDates() {
    const closingDates = [];
    let rowIndex = 1;
    
    while (true) {
        const xpath = `/html/body/main/div[1]/div[1]/div/div/div/div[1]/div[2]/div/div/form/table/tbody/tr[${rowIndex}]/td/div/span/span/span[2]/span[2]`;
        try {
            const result = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            if (result.singleNodeValue) {
                const date = result.singleNodeValue.textContent.trim();
                closingDates.push({
                    rowNumber: rowIndex,
                    closingDate: date
                });
                rowIndex++;
            } else {
                break;
            }
        } catch (e) {
            break;
        }
    }
    return closingDates;
}
return getAllClosingDates();
"""

GET_CATEGORY_SCRIPT = """
function getCategory() {
    // First click the Categories tab
    const tabSelectors = {
        css: '#categoriesAbstractTab > a',
        xpath: '//*[@id="categoriesAbstractTab"]/a',
        fullXpath: '/html/body/main/div[1]/div[1]/div[1]/ul/li[2]/a'
    };

    // Click the tab
    let tabClicked = false;
    for (let selectorType in tabSelectors) {
        try {
            let element;
            const selector = tabSelectors[selectorType];
            
            if (selectorType === 'css') {
                element = document.querySelector(selector);
            } else {
                const result = document.evaluate(selector, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
                element = result.singleNodeValue;
            }

            if (element) {
                element.click();
                tabClicked = true;
                break;
            }
        } catch (e) {}
    }

    if (!tabClicked) return null;

    // Wait for content to load
    return new Promise(resolve => {
        setTimeout(() => {
            try {
                const codeElement = document.querySelector('span[id^="selectedCategoryContainerNIGP_SS"][id$="-code"]');
                const nameElement = document.querySelector('span[id^="selectedCategoryContainerNIGP_SS"][id$="-name"]');
                if (codeElement && nameElement) {
                    resolve({
                        code: codeElement.textContent.trim(),
                        name: nameElement.textContent.trim()
                    });
                }
                resolve(null);
            } catch (e) {
                resolve(null);
            }
        }, 2000);
    });
}
return getCategory();
"""

from utils.utils import safe_move, play_notification_sound

import time

import random
import logging
import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Set
from dataclasses import dataclass, asdict
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
from openpyxl.utils import get_column_letter
import re


# List of common user agents
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
]


# Load environment variables
load_dotenv()
BIDNET_EMAIL = os.getenv("BIDNET_EMAIL")
BIDNET_PASSWORD = os.getenv("BIDNET_PASSWORD")


@dataclass
class BidData:
    """Data structure for storing bid information"""

    sl_no: int
    posted_date: str
    response_date: str
    notice_type: str
    solicitation_number: str
    solicitation_title: str
    agency: str
    category: str
    description: str
    additional_summary: str
    contracting_office_address: str
    contact_information: str
    bid_detail_page_url: str
    attachments: str


class BidNetScraper:
    def __init__(self, days: int = 2):
        """Initialize the BidNet scraper with configuration"""
        self.days = days
        self.base_url = "https://www.bidnetdirect.com"
        self.driver = None
        self.processed_bids: Set[str] = set()
        self.current_date = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        self.bids_data = []
        self.cookie_banner_handled = False  # Track cookie banner state

        # Update folder structure
        self.script_name = "17_BidNet"
        self.main_folder = Path(self.current_date)
        self.script_folder_in_progress = self.main_folder / f"{self.script_name}_IN_PROGRESS"
        self.script_folder = self.script_folder_in_progress / self.script_name
        
        # Update cache location and ensure it exists
        self.cache_dir = Path("cache")
        self.cache_dir.mkdir(exist_ok=True)
        self.cache_file = self.cache_dir / f"{self.script_name}_cache.json"
        
        # Create empty cache file if it doesn't exist
        if not self.cache_file.exists():
            with open(self.cache_file, "w") as f:
                json.dump({}, f, indent=2)
        
        self.log_file = self.main_folder / "bidnet_scraper.log"

        # Add JavaScript scripts as instance variables
        self.GET_BID_DETAILS_SCRIPT = GET_BID_DETAILS_SCRIPT
        self.GET_CATEGORY_SCRIPT = GET_CATEGORY_SCRIPT

        # Setup logging after folder structure is defined
        self.setup_logging()


    def setup_logging(self):
        """Configure logging with proper directory structure"""
        # Configure logging
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)

        # Clear any existing handlers
        logger.handlers.clear()

        # Create formatter
        formatter = logging.Formatter("%(message)s")

        # Console handler only
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

        # Store logger as instance variable
        self.logger = logger

    def setup_folders(self):
        """Create necessary folders for downloads and processing"""
        try:
            # Create main folder if it doesn't exist
            self.main_folder.mkdir(parents=True, exist_ok=True)
            
            # Create IN_PROGRESS folder only if it doesn't exist
            if not self.script_folder_in_progress.exists():
                self.script_folder_in_progress.mkdir(parents=True)
                self.logger.info(f"Created processing folder: {self.script_folder_in_progress}")
            
            # Create script folder inside IN_PROGRESS
            self.script_folder.mkdir(parents=True, exist_ok=True)
            
            # Create cache file parent directory
            self.cache_file.parent.mkdir(parents=True, exist_ok=True)
            
        except Exception as e:
            self.logger.error(f"Error setting up folders: {str(e)}")
            raise

    def setup_driver(self):
        """Setup Selenium WebDriver with anti-bot measures"""
        options = webdriver.ChromeOptions()
        options.add_argument(f"user-agent={random.choice(USER_AGENTS)}")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)

        # Ensure download directory exists before setting it
        self.script_folder.mkdir(parents=True, exist_ok=True)
        
        prefs = {
            "download.default_directory": str(self.script_folder.absolute()),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)

        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()

        self.driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {
                "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                })
                """
            },
        )

    def random_delay(self, min_seconds: float = 1.0, max_seconds: float = 3.0):
        """Add random delay between actions"""
        time.sleep(random.uniform(min_seconds, max_seconds))

    def simulate_human_input(self, element, text: str):
        """Simulate human-like typing"""
        for char in text:
            element.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))

    def login(self) -> bool:
        """Handle login process with SAML authentication and enhanced error handling"""
        try:
            self.logger.info("Starting login process...")
            max_retries = 3
            retry_count = 0

            while retry_count < max_retries:
                try:
                    # Start with SAML login URL directly
                    self.driver.get("https://www.bidnetdirect.com/saml/login")
                    self.random_delay(5, 7)  # Increased initial wait

                    # Log current URL for debugging
                    current_url = self.driver.current_url
                    self.logger.info(f"Current URL after navigation: {current_url}")

                    # Additional check for page load
                    try:
                        WebDriverWait(self.driver, 10).until(
                            lambda driver: driver.execute_script(
                                "return document.readyState"
                            )
                            == "complete"
                        )
                    except Exception as e:
                        self.logger.error(f"Page load timeout: {str(e)}")
                        retry_count += 1
                        continue

                    # Wait for page load with multiple selectors
                    username_selectors = [
                        (By.ID, "j_username"),
                        (By.NAME, "j_username"),
                        (By.CSS_SELECTOR, "input[id='j_username']"),
                        (By.CSS_SELECTOR, "input[name='j_username']"),
                        (By.XPATH, "//input[@id='j_username']"),
                        (By.XPATH, "//input[@name='j_username']"),
                        (By.XPATH, "//input[@type='text']"),
                    ]

                    username = None
                    for selector_type, selector_value in username_selectors:
                        try:
                            self.logger.debug(
                                f"Trying to find username field with {selector_type}: {selector_value}"
                            )
                            username = WebDriverWait(
                                self.driver, 10
                            ).until(EC.presence_of_element_located((selector_type, selector_value)))
                            if (
                                username
                                and username.is_displayed()
                                and username.is_enabled()
                            ):
                                self.logger.info(
                                    f"Found username field using {selector_type}: {selector_value}"
                                )
                                break
                        except Exception as e:
                            self.logger.debug(
                                f"Selector {selector_value} failed: {str(e)}"
                            )
                            continue

                    if not username:
                        self.logger.error(
                            f"Could not find username field with any selector (attempt {retry_count + 1}/{max_retries})"
                        )
                        retry_count += 1
                        self.random_delay(3, 5)
                        continue

                    # Additional verification that we're on the right page
                    if "idp.bidnetdirect.com" not in self.driver.current_url:
                        self.logger.error(
                            f"Not on expected login page. Current URL: {self.driver.current_url}"
                        )
                        retry_count += 1
                        continue

                    # Clear any existing input and type username
                    username.clear()
                    self.random_delay(0.5, 1)
                    self.simulate_human_input(username, BIDNET_EMAIL)
                    self.random_delay(1, 2)

                    # Find and fill password with similar approach
                    password_selectors = [
                        (By.ID, "j_password"),
                        (By.NAME, "j_password"),
                        (By.CSS_SELECTOR, "input[id='j_password']"),
                        (By.CSS_SELECTOR, "input[name='j_password']"),
                        (By.XPATH, "//input[@id='j_password']"),
                        (By.XPATH, "//input[@name='j_password']"),
                        (By.XPATH, "//input[@type='password']"),
                    ]

                    password = None
                    for selector_type, selector_value in password_selectors:
                        try:
                            password = WebDriverWait(
                                self.driver, 10
                            ).until(EC.presence_of_element_located((selector_type, selector_value)))
                            if (
                                password
                                and password.is_displayed()
                                and password.is_enabled()
                            ):
                                break
                        except Exception as e:
                            self.logger.debug(
                                f"Password selector {selector_value} failed: {str(e)}"
                            )
                            continue

                    if not password:
                        self.logger.error(
                            f"Could not find password field (attempt {retry_count + 1}/{max_retries})"
                        )
                        retry_count += 1
                        self.random_delay(3, 5)
                        continue

                    password.clear()
                    self.random_delay(0.5, 1)
                    self.simulate_human_input(password, BIDNET_PASSWORD)
                    self.random_delay(1, 2)

                    # Find and click login button with multiple selectors
                    login_button_selectors = [
                        (By.ID, "loginButton"),
                        (By.CSS_SELECTOR, "button[type='submit']"),
                        (By.XPATH, "//button[@type='submit']"),
                        (By.XPATH, "//button[contains(text(), 'Login')]"),
                        (By.XPATH, "//input[@type='submit']"),
                    ]

                    login_button = None
                    for selector_type, selector_value in login_button_selectors:
                        try:
                            login_button = WebDriverWait(
                                self.driver, 10
                            ).until(EC.element_to_be_clickable((selector_type, selector_value)))
                            if (
                                login_button
                                and login_button.is_displayed()
                                and login_button.is_enabled()
                            ):
                                break
                        except Exception as e:
                            self.logger.debug(
                                f"Login button selector {selector_value} failed: {str(e)}"
                            )
                            continue

                    if not login_button:
                        self.logger.error(
                            f"Could not find login button (attempt {retry_count + 1}/{max_retries})"
                        )
                        retry_count += 1
                        self.random_delay(3, 5)
                        continue

                    # Move mouse naturally to button and click
                    try:
                        ActionChains(self.driver).move_to_element(login_button).pause(
                            1
                        ).click().perform()
                    except Exception as e:
                        self.logger.error(f"Failed to click login button: {str(e)}")
                        retry_count += 1
                        continue

                    # Wait longer for SAML redirect and processing
                    self.random_delay(6, 8)  # Increased wait time

                    # Multiple success checks with detailed logging
                    verification_retries = 3
                    verification_count = 0
                    while verification_count < verification_retries:
                        try:
                            self.logger.info(
                                f"Login verification attempt {verification_count + 1}"
                            )

                            # Check if we're on the expected page after login
                            if (
                                "private/supplier/solicitations/search"
                                in self.driver.current_url
                            ):
                                self.logger.info("Login successful - correct URL")
                                # Add cookie banner handling after successful login
                                self.random_delay(2, 3)  # Wait for cookie banner to appear
                                try:
                                    cookie_script = """
                                    function handleCookieBanner() {
                                        const button = document.querySelector('#cookieBannerAcceptBtn');
                                        if (button) {
                                            button.click();
                                            return true;
                                        }
                                        return false;
                                    }
                                    return handleCookieBanner();
                                    """
                                    banner_handled = self.driver.execute_script(cookie_script)
                                    if banner_handled:
                                        self.logger.info("Cookie banner accepted")
                                        time.sleep(2)  # Wait for banner to disappear
                                except Exception as e:
                                    self.logger.error(f"Error handling cookie banner: {str(e)}")
                                return True

                            # Additional check for any error messages
                            error_messages = self.driver.find_elements(
                                By.CLASS_NAME, "error-message"
                            )
                            if error_messages:
                                self.logger.error(
                                    f"Login error message found: {error_messages[0].text}"
                                )
                                break  # Break inner loop to retry login

                            # If we're still on login page, login failed
                            if "login" in self.driver.current_url.lower():
                                self.logger.error("Still on login page - login failed")
                                break  # Break inner loop to retry login

                            verification_count += 1
                            self.random_delay(2, 3)

                        except Exception as e:
                            self.logger.error(
                                f"Error during login verification: {str(e)}"
                            )
                            verification_count += 1
                            self.random_delay(2, 3)

                    retry_count += 1

                except Exception as e:
                    self.logger.error(
                        f"Login attempt {retry_count + 1} failed: {str(e)}"
                    )
                    retry_count += 1
                    self.random_delay(3, 5)

            self.logger.error("Login failed after all retries")
            return False

        except Exception as e:
            self.logger.error(f"Critical login error: {str(e)}")
            return False

    def load_cache(self) -> Dict:
        """Load and clean previously processed bids from cache"""
        try:
            if self.cache_file.exists():
                with open(self.cache_file, "r") as f:
                    cache = json.load(f)

                # Clean old entries (older than 3 months)
                current_date = datetime.now()
                cleaned_cache = {}

                for bid_url, data in cache.items():
                    try:
                        posted_date = datetime.strptime(data["posted_date"], "%Y-%m-%d")
                        if (current_date - posted_date).days <= 90:  # Keep bids from last 3 months
                            cleaned_cache[bid_url] = data
                    except Exception as e:
                        self.logger.error(f"Error processing cache entry: {str(e)}")
                        continue

                # Save cleaned cache
                with open(self.cache_file, "w") as f:
                    json.dump(cleaned_cache, f, indent=2)

                return cleaned_cache
            return {}
        except Exception as e:
            self.logger.error(f"Error loading cache: {str(e)}")
            return {}

    def save_to_cache(self, bid_data: Dict):
        """Save processed bid to cache with metadata"""
        try:
            cache = self.load_cache()
            
            # Add debug logging
            self.logger.info(f"Pre-conversion posted date: {bid_data['posted_date']}")
            
            try:
                if bid_data["posted_date"]:
                    # Check if date is already in YYYY-MM-DD format
                    if re.match(r'^\d{4}-\d{2}-\d{2}$', bid_data["posted_date"]):
                        formatted_date = bid_data["posted_date"]
                        self.logger.info(f"Date already in correct format: {formatted_date}")
                    else:
                        # Convert from MM/DD/YYYY to YYYY-MM-DD
                        posted_date = datetime.strptime(bid_data["posted_date"], "%m/%d/%Y")
                        formatted_date = posted_date.strftime("%Y-%m-%d")
                        self.logger.info(f"Converted date format: {formatted_date}")
                else:
                    formatted_date = datetime.now().strftime("%Y-%m-%d")
                    self.logger.info(f"Using current date: {formatted_date}")
                    
                cache[bid_data["bid_detail_page_url"]] = {
                    "posted_date": formatted_date,
                    "last_checked": datetime.now().strftime("%Y-%m-%d"),
                    "solicitation_number": bid_data["solicitation_number"]
                }
                
                with open(self.cache_file, "w") as f:
                    json.dump(cache, f, indent=2)
                self.logger.info(f"✅ Successfully cached bid: {bid_data['solicitation_number']}")
                
            except ValueError as e:
                self.logger.error(f"Error formatting date for cache: {str(e)}")
                self.logger.error(f"Problematic date value: {bid_data['posted_date']}")
                # Use original date if conversion fails
                cache[bid_data["bid_detail_page_url"]] = {
                    "posted_date": bid_data["posted_date"],
                    "last_checked": datetime.now().strftime("%Y-%m-%d"),
                    "solicitation_number": bid_data["solicitation_number"]
                }
                with open(self.cache_file, "w") as f:
                    json.dump(cache, f, indent=2)
                self.logger.info(f"✅ Cached bid with original date format: {bid_data['solicitation_number']}")
                
        except Exception as e:
            self.logger.error(f"Error saving to cache: {str(e)}")

    def is_bid_in_cache(self, url: str, solicitation_number: str) -> bool:
        """Check if bid is already in cache"""
        try:
            cache = self.load_cache()
            
            # Check by URL
            if url in cache:
                self.logger.info(f"⏭️ Bid already processed (URL): {solicitation_number}")
                return True
            
            # Also check by solicitation number as fallback
            for entry in cache.values():
                if entry.get("solicitation_number") == solicitation_number:
                    self.logger.info(f"⏭️ Bid already processed (Number): {solicitation_number}")
                    return True
                
            return False
            
        except Exception as e:
            self.logger.error(f"Error checking cache: {str(e)}")
            return False

    def check_incomplete_form(self) -> bool:
        """Check if there's an incomplete purchasing group form"""
        try:
            # Check using multiple selectors
            check_script = """
                function checkIncompleteForm() {
                    // Check by ID
                    const titleElement = document.querySelector('#ui-id-1');
                    if (titleElement && titleElement.textContent.includes('Incomplete Purchasing Group Form')) {
                        return true;
                    }
                    
                    // Check by XPath as backup
                    const xpath = '/html/body/div[11]/div[1]/span';
                    const result = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
                    const xpathElement = result.singleNodeValue;
                    if (xpathElement && xpathElement.textContent.includes('Incomplete Purchasing Group Form')) {
                        return true;
                    }
                    
                    return false;
                }
                return checkIncompleteForm();
            """
            is_incomplete = self.driver.execute_script(check_script)
            if is_incomplete:
                self.logger.warning("⚠️ Incomplete Purchasing Group Form detected - skipping")
                return True
            return False
            
        except Exception as e:
            self.logger.error(f"Error checking incomplete form: {str(e)}")
            return False

    def process_bid_links(self, links: List[Dict[str, str]]) -> None:
        """Process list of bid links and save to Excel after each bid"""
        self.logger.info(f"Processing {len(links)} bid links")

        for link in links:
            try:
                self.logger.info(f"\nProcessing bid: {link['title']}")

                # Navigate to bid page
                self.driver.get(link["url"])
                self.random_delay(2, 3)

                # Check for incomplete form before proceeding
                if self.check_incomplete_form():
                    continue

                # Extract bid details with dates from link
                bid_data = self.extract_bid_details(
                    url=link["url"],
                    posted_date=link["publicationDate"],
                    response_date=link["closingDate"]
                )
                if not bid_data:
                    continue

                # Check if bid is already processed
                if self.is_bid_in_cache(link["url"], bid_data.solicitation_number):
                    continue

                # Create folders only after we know we'll process this bid
                clean_solicitation = re.sub(r'[<>:"/\\|?*]', '_', bid_data.solicitation_number)
                bid_folder = self.script_folder_in_progress / clean_solicitation
                bid_folder.mkdir(parents=True, exist_ok=True)

                # Download attachments
                attachments = self.download_bid_attachments(bid_data.solicitation_number)
                bid_data.attachments = attachments
                if attachments:
                    self.logger.info(f"Downloaded Attachments: {attachments}")

                # Format dates before saving
                bid_data.posted_date = self.format_date(bid_data.posted_date)
                bid_data.response_date = self.format_date(bid_data.response_date)

                # Save to cache
                self.save_to_cache(asdict(bid_data))

                # Update Excel file
                if self.update_excel_after_bid(bid_data):
                    self.logger.info(f"✅ Successfully processed and saved bid: {bid_data.solicitation_number}")
                else:
                    self.logger.error(f"❌ Failed to save bid to Excel: {bid_data.solicitation_number}")

                self.random_delay(2, 4)

            except Exception as e:
                self.logger.error(f"❌ Error processing bid {link['title']}: {str(e)}")
                play_notification_sound()
                input("Press Enter to continue...")

    def __enter__(self):
        """Context manager entry"""
        try:
            self.setup_folders()
            self.setup_driver()
            return self

        except Exception as e:
            self.logger.error(f"Error in setup: {str(e)}")
            if self.driver:
                self.driver.quit()
            raise

    def mark_processing_complete(self):
        """Change folder suffix from _IN_PROGRESS to _COMPLETED"""
        try:
            completed_folder = self.main_folder / f"{self.script_name}_COMPLETED"
            
            # Clean up temporary download folder first
            try:
                if self.script_folder.exists():
                    shutil.rmtree(self.script_folder, ignore_errors=True)
                    print(f"✅ Removed temporary download folder: {self.script_folder}")
            except Exception as e:
                print(f"⚠️ Error removing temporary folder: {str(e)}")
            
            # If COMPLETED folder already exists, remove it first
            if completed_folder.exists():
                import shutil
                shutil.rmtree(completed_folder)
                self.logger.info(f"Removed existing completed folder: {completed_folder}")
            
            # Only attempt rename if IN_PROGRESS folder exists
            if self.script_folder_in_progress.exists():
                self.script_folder_in_progress.rename(completed_folder)
                print(f"✅ Successfully renamed processing folder to {completed_folder}")
            else:
                print("⚠️ Processing folder does not exist, skipping rename")
                
        except Exception as e:
            print(f"❌ Error renaming processing folder: {str(e)}")

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit with folder status update"""
        try:
            if self.driver:
                self.driver.quit()
            self.mark_processing_complete()
        except Exception as e:
            print(f"❌ Error during cleanup: {str(e)}")
        """Context manager exit"""
        if self.driver:
            self.driver.quit()

    def navigate_to_purchasing_groups(self) -> bool:
        """Navigate to the purchasing groups page"""
        try:
            self.driver.get(f"{self.base_url}/purchasing-groups")
            self.random_delay(2, 4)
            return True
        except Exception as e:
            self.logger.error(f"Failed to navigate to purchasing groups: {str(e)}")
            return False

    def get_purchasing_groups(self) -> List[Dict[str, str]]:
        """Extract all purchasing groups from the page"""
        groups = []
        try:
            group_elements = self.driver.find_elements(
                By.CSS_SELECTOR,
                "#main > div > div > div > section > div > div > div.leftCol > div > ul > li > a",
            )
            for element in group_elements:
                groups.append(
                    {"name": element.text, "url": element.get_attribute("href")}
                )
            self.logger.info(f"Found {len(groups)} purchasing groups")
            return groups
        except Exception as e:
            self.logger.error(f"Failed to extract purchasing groups: {str(e)}")
            return []

    def handle_cookie_banner(self):
        """Handle cookie banner if present"""
        try:
            if self.cookie_banner_handled:
                return True
                
            # Try to find and click the cookie accept button
            cookie_script = """
            function handleCookieBanner() {
                const button = document.querySelector('#cookieBannerAcceptBtn');
                if (button && button.offsetParent !== null) {  // Check if button is visible
                    button.click();
                    return true;
                }
                return false;
            }
            return handleCookieBanner();
            """
            banner_handled = self.driver.execute_script(cookie_script)
            if banner_handled:
                self.logger.info("Cookie banner accepted")
                self.cookie_banner_handled = True  # Set flag when banner is handled
                time.sleep(2)  # Wait for banner to disappear
            return banner_handled
        except Exception as e:
            self.logger.error(f"Error handling cookie banner: {str(e)}")
            return False

    def check_group_bids_tab(self) -> Optional[int]:
        """Check if group bids tab exists and return bid count"""
        try:
            # Handle cookie banner only if not already handled
            if not self.cookie_banner_handled:
                self.handle_cookie_banner()
            
            # First check if the button exists without waiting
            group_bids_button = self.driver.find_elements(By.ID, "selectedContent-buyer")
            if not group_bids_button:
                self.logger.info("⏭️ No Group Bids tab found for this group")
                return None
                
            # Now that we know the button exists, wait for it to be clickable
            group_bids_button = WebDriverWait(self.driver, 20).until(
                EC.element_to_be_clickable((By.ID, "selectedContent-buyer"))
            )
            
            # Scroll into view and ensure no overlays
            self.driver.execute_script("arguments[0].scrollIntoView(true);", group_bids_button)
            self.random_delay(1, 2)
            
            # Click using JavaScript and wait for count to update
            self.driver.execute_script("arguments[0].click();", group_bids_button)
            self.random_delay(2, 3)
            
            # Get bid count after clicking the button
            count_element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#selectedContent-buyer .filter-count"))
            )
            count_text = count_element.text.strip()
            
            if count_text.isdigit():
                return int(count_text)
            
            return None
                
        except Exception as e:
            self.logger.error(f"Error checking group bids tab: {str(e)}")
            return None


    def apply_sort_order_descending(self) -> bool:
        """Apply descending sort order for bids"""
        try:
            sort_script = """
            function setSortOrder() {
                const selectWrapper = document.querySelector('.select-wrapper');
                if (!selectWrapper) return false;
                
                const dropdown = selectWrapper.querySelector('.select-dropdown');
                if (!dropdown) return false;
                
                dropdown.click();
                
                const options = document.querySelectorAll('.select-dropdown li');
                for (let option of options) {
                    if (option.textContent.trim() === "Start Date (Descending)") {
                        option.click();
                        return true;
                    }
                }
                return false;
            }
            return setSortOrder();
            """
            result = self.driver.execute_script(sort_script)
            if result:
                self.random_delay(2, 3)  # Wait for sort to take effect
                return True
            return False
        except Exception as e:
            return False

    def extract_bid_links(self) -> List[Dict[str, str]]:
        """Extract bid links using JavaScript execution"""
        try:
            # Apply descending sort order first
            if not self.apply_sort_order_descending():
                print("Continuing with default order")

            # Get all bid details using JavaScript
            script = """
            function getBidLinks() {
                const bids = [];
                const links = document.querySelectorAll('a.solicitation-link');
                
                for (const link of links) {
                    try {
                        const row = link.closest('tr');
                        if (!row) continue;
                        
                        const publicationDate = row.querySelector('span.sol-publication-date span.date-value')?.textContent.trim();
                        const closingDate = row.querySelector('span.sol-closing-date span.date-value')?.textContent.trim();
                        
                        // Construct full URL from href
                        const href = link.getAttribute('href');
                        const fullUrl = href.startsWith('http') ? href : 'https://www.bidnetdirect.com' + href;
                        
                        bids.push({
                            title: link.textContent.trim(),
                            url: fullUrl,
                            publicationDate: publicationDate,
                            closingDate: closingDate
                        });
                    } catch (e) {
                        console.error('Error processing bid link:', e);
                    }
                }
                return bids;
            }
            return getBidLinks();
            """
            
            bids = self.driver.execute_script(script)
            
            # Filter bids by date range
            filtered_bids = []
            for bid in bids:
                if self.is_within_date_range(bid["publicationDate"]):
                    filtered_bids.append(bid)
                else:
                    # Since bids are sorted by date, we can break early
                    break
                    
            self.logger.info(f"Found {len(filtered_bids)} bids within date range")
            return filtered_bids
        except Exception as e:
            self.logger.error(f"Failed to extract bid links: {str(e)}")
            return []


    def get_publication_dates(self) -> List[Dict[str, str]]:
        """Get all publication dates using JavaScript"""
        script = """
        function getAllPublicationDates() {
            const publicationDates = [];
            let rowIndex = 1;
            
            while (true) {
                const xpath = `/html/body/main/div[1]/div[1]/div/div/div/div[1]/div[2]/div/div/form/table/tbody/tr[${rowIndex}]/td/div/span/span/span[1]/span[2]`;
                
                try {
                    const result = document.evaluate(
                        xpath,
                        document,
                        null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE,
                        null
                    );
                    
                    if (result.singleNodeValue) {
                        const date = result.singleNodeValue.textContent.trim();
                        publicationDates.push({
                            rowNumber: rowIndex,
                            publicationDate: date
                        });
                        rowIndex++;
                    } else {
                        break;
                    }
                } catch (e) {
                    break;
                }
            }
            return publicationDates;
        }
        return getAllPublicationDates();
        """
        try:
            dates = self.driver.execute_script(script)
            self.logger.info(f"Found {len(dates)} publication dates")
            return dates
        except Exception as e:
            self.logger.error(f"Failed to get publication dates: {str(e)}")
            return []


    def get_closing_dates(self) -> List[Dict[str, str]]:
        """Get all closing dates using JavaScript"""
        script = """
        function getAllClosingDates() {
            const closingDates = [];
            let rowIndex = 1;
            
            while (true) {
                const xpath = `/html/body/main/div[1]/div[1]/div/div/div/div[1]/div[2]/div/div/form/table/tbody/tr[${rowIndex}]/td/div/span/span/span[2]/span[2]`;
                
                try {
                    const result = document.evaluate(
                        xpath,
                        document,
                        null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE,
                        null
                    );
                    
                    if (result.singleNodeValue) {
                        const date = result.singleNodeValue.textContent.trim();
                        closingDates.push({
                            rowNumber: rowIndex,
                            closingDate: date
                        });
                        rowIndex++;
                    } else {
                        break;
                    }
                } catch (e) {
                    break;
                }
            }
            return closingDates;
        }
        return getAllClosingDates();
        """
        try:
            dates = self.driver.execute_script(script)
            self.logger.info(f"Found {len(dates)} closing dates")
            return dates
        except Exception as e:
            self.logger.error(f"Failed to get closing dates: {str(e)}")
            return []


    def get_bid_details_js(self) -> List[Dict[str, str]]:
        """Get all bid details using JavaScript"""
        script = """
        function getBidDetails() {
            const bids = [];
            const rows = document.querySelectorAll('table tbody tr');
            
            rows.forEach((row, index) => {
                try {
                    const titleLink = row.querySelector('a.solicitation-link');
                    if (!titleLink) return;
                    
                    const bid = {
                        rowNumber: index + 1,
                        title: titleLink.textContent.trim(),
                        url: titleLink.href,
                        publicationDate: row.querySelector('span.sol-publication-date span.date-value')?.textContent.trim(),
                        closingDate: row.querySelector('span.sol-closing-date span.date-value')?.textContent.trim()
                    };
                    
                    bids.push(bid);
                } catch (e) {
                    console.error(`Error processing row ${index + 1}:`, e);
                }
            });
            
            return bids;
        }
        return getBidDetails();
        """
        try:
            bids = self.driver.execute_script(script)
            self.logger.info(f"Found {len(bids)} bids")
            return bids
        except Exception as e:
            self.logger.error(f"Failed to get bid details: {str(e)}")
            return []


    def is_within_date_range(self, date_str: str) -> bool:
        """Check if the bid date is within the specified range"""
        try:
            bid_date = datetime.strptime(date_str, "%m/%d/%Y")
            cutoff_date = datetime.now() - timedelta(days=self.days)
            return bid_date >= cutoff_date
        except Exception as e:
            self.logger.error(f"Date parsing error: {str(e)}")
            return False


    def extract_contact_info(
        self,
        name: str = "",
        title: str = "",
        phone: str = "",
        fax: str = "",
        email: str = "",
    ) -> str:
        """Format contact information consistently"""
        contact_parts = []
        if name:
            contact_parts.append(f"Name: {name}")
        if title:
            contact_parts.append(f"Title: {title}")
        if phone:
            contact_parts.append(f"Phone: {phone}")
        if fax:
            contact_parts.append(f"Fax: {fax}")
        if email:
            contact_parts.append(f"Email: {email}")
        return "\n".join(contact_parts)


    def _safe_get_element_text(self, selector: str, timeout: int = 5) -> str:
        """Safely get element text with timeout"""
        try:
            element = WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            return element.text
        except:
            return ""


    def _safe_get_element_text_with_multiple_methods(
        self, selectors: List[Dict[str, str]], timeout: int = 10
    ) -> Optional[str]:
        """
        Try multiple methods to get element text
        selectors: List of dicts with 'type' (css/xpath) and 'value'
        """
        for selector in selectors:
            try:
                if selector["type"] == "css":
                    element = WebDriverWait(self.driver, timeout).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector["value"]))
                    )
                else:  # xpath
                    element = WebDriverWait(self.driver, timeout).until(
                        EC.presence_of_element_located((By.XPATH, selector["value"]))
                    )
                if element and element.text.strip():
                    return element.text.strip()
            except Exception as e:
                self.logger.debug(
                    f"Failed to find element using {selector['type']}: {selector['value']}"
                )
                continue
        return None


    def extract_category_info_with_fallback(self) -> Optional[str]:
        """Extract category information with multiple fallback strategies"""
        try:
            # Try clicking the Categories tab
            tab_clicked = False
            selectors = [
                (By.CSS_SELECTOR, "#categoriesAbstractTab > a"),
                (By.XPATH, "//a[contains(text(), 'Categories')]"),
                (By.PARTIAL_LINK_TEXT, "Categories"),
            ]

            for selector_type, selector in selectors:
                try:
                    tab = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((selector_type, selector))
                    )
                    tab.click()
                    tab_clicked = True
                    self.random_delay(2, 3)
                    break
                except Exception:
                    continue

            if not tab_clicked:
                return None

            # Try JavaScript extraction first
            try:
                category_info = self.driver.execute_script(
                    """
                    const codeElement = document.querySelector('span[id^="selectedCategoryContainerNIGP_SS"][id$="-code"]');
                    const nameElement = document.querySelector('span[id^="selectedCategoryContainerNIGP_SS"][id$="-name"]');
                    if (codeElement && nameElement) {
                        return `${codeElement.textContent.trim()} - ${nameElement.textContent.trim()}`;
                    }
                    return null;
                """
                )
                if category_info:
                    return category_info
            except Exception:
                pass

            # Fallback to direct element extraction
            try:
                code_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located(
                        (
                            By.CSS_SELECTOR,
                            "span[id^='selectedCategoryContainerNIGP_SS'][id$='-code']",
                        )
                    )
                )
                if code_elements:
                    code_element = code_elements[0]
                    name_element = self.driver.find_element(
                        By.CSS_SELECTOR,
                        f"#{code_element.get_attribute('id').replace('-code', '-name')}",
                    )
                    return f"{code_element.text.strip()} - {name_element.text.strip()}"
            except Exception:
                pass

            return None

        except Exception as e:
            self.logger.error(f"Failed to extract category information: {str(e)}")
            return None


    def extract_bid_details(self, url: str, posted_date: str, response_date: str) -> Optional[BidData]:
        """Extract bid details using JavaScript execution"""
        try:
            self.logger.info(f"📄 Extracting details from bid: {url}")
            self.driver.get(url)
            self.random_delay(3, 5)

            # Get main bid details using JavaScript
            details_script = """
            function getBidDetails() {
                // Check for Owner Organization
                const ownerOrgXPath = '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[3]/span';
                const ownerOrgElement = document.evaluate(ownerOrgXPath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                const isOwnerOrgPresent = ownerOrgElement && ownerOrgElement.textContent.trim() === 'Owner Organization';
                
                console.log("Owner Organization present:", isOwnerOrgPresent);

                // Define XPaths based on Owner Organization presence
                const xpaths = {
                    noticeType: isOwnerOrgPresent ? 
                        '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[4]/div/p' : 
                        '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[3]/div/p',
                    solNumber: isOwnerOrgPresent ? 
                        '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[5]/div/p' : 
                        '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[4]/div/p',
                    title: isOwnerOrgPresent ? 
                        '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[6]/div/p' : 
                        '/html/body/main/div[1]/div[2]/div[3]/form/div[1]/div[1]/div[5]/div/p'
                };

                function getElementText(xpath) {
                    const element = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                    return element ? element.textContent.trim() : '';
                }

                // Get organization name
                const agency = document.querySelector('#g_548 > div > p > a')?.textContent.trim() || '';

                // Get description
                const description = document.querySelector('#descriptionText')?.textContent.trim() || '';

                // Get additional instructions
                let additionalSummary = '';
                const instructionsElement = document.querySelector('#tenderInstructionsEN_0');
                if (instructionsElement) {
                    try {
                        const readMoreLink = instructionsElement.querySelector('.read-more');
                        if (readMoreLink) readMoreLink.click();
                    } catch (e) {}
                    additionalSummary = instructionsElement.textContent.replace('See more', '').trim();
                }

                // Get contracting office address
                const address = document.querySelector('#g_632 > div > p')?.textContent.trim() || '';

                // Get contact information
                const contactInfo = {
                    name: document.querySelector('#g_639 > div > p')?.textContent.trim() || '',
                    title: 'Contracting Officer',
                    phone: document.querySelector('#g_640 > div > p')?.textContent.trim() || '',
                    email: document.querySelector('#g_641 > div > p')?.textContent.trim() || ''
                };

                return {
                    notice_type: getElementText(xpaths.noticeType),
                    solicitation_number: getElementText(xpaths.solNumber),
                    title: getElementText(xpaths.title),
                    agency: agency,
                    description: description,
                    additional_summary: additionalSummary,
                    contracting_office_address: address,
                    contact_info: contactInfo
                };
            }
            return getBidDetails();
            """

            details = self.driver.execute_script(details_script)
            if not details:
                self.logger.error("Failed to extract bid details")
                return None

            # Format contact information
            contact_info = details["contact_info"]
            contact_parts = []
            if contact_info["name"]:
                contact_parts.append(f"Name: {contact_info['name']}")
            if contact_info["title"]:
                contact_parts.append(f"Title: {contact_info['title']}")
            if contact_info["phone"]:
                contact_parts.append(f"Phone: {contact_info['phone']}")
            if contact_info["email"]:
                contact_parts.append(f"Email: {contact_info['email']}")
            contact_str = "\n".join(contact_parts)

            # Get category info
            category_str = ""
            try:
                categories_tab = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#categoriesAbstractTab > a"))
                )
                categories_tab.click()
                self.random_delay(2, 3)

                category_script = """
                function getCategory() {
                    const code = document.querySelector('span[id^="selectedCategoryContainerNIGP_SS"][id$="-code"]')?.textContent.trim() || '';
                    const name = document.querySelector('span[id^="selectedCategoryContainerNIGP_SS"][id$="-name"]')?.textContent.trim() || '';
                    return code && name ? `${code} - ${name}` : '';
                }
                return getCategory();
                """
                category_str = self.driver.execute_script(category_script)
            except:
                self.logger.info("Category info not available")
                category_str = ""

            # Create bid data object
            bid_data = BidData(
                sl_no=len(self.bids_data) + 1,
                posted_date=posted_date,
                response_date=response_date,
                notice_type=details["notice_type"],
                solicitation_number=details["solicitation_number"],
                solicitation_title=details["title"],
                agency=details["agency"],
                category=category_str,
                description=details["description"],
                additional_summary=details["additional_summary"],
                contracting_office_address=details["contracting_office_address"],
                contact_information=contact_str,
                bid_detail_page_url=url,
                attachments=""
            )

            # Log extracted details
            self.logger.info("\nExtracted Bid Details:")
            self.logger.info(f"Solicitation Number: {bid_data.solicitation_number}")
            self.logger.info(f"Title: {bid_data.solicitation_title}")
            self.logger.info(f"Agency: {bid_data.agency}")
            self.logger.info(f"Posted Date: {bid_data.posted_date}")
            self.logger.info(f"Response Date: {bid_data.response_date}")
            self.logger.info(f"Notice Type: {bid_data.notice_type}")
            self.logger.info(f"Category: {bid_data.category}")

            return bid_data

        except Exception as e:
            self.logger.error(f"Error extracting bid details: {str(e)}")
            return None


    def download_bid_attachments(self, solicitation_number: str) -> str:
        """Download attachments for a bid and return comma-separated list of filenames"""
        try:
            if not solicitation_number:
                self.logger.info("No solicitation number provided")
                return ""
                
            clean_solicitation = re.sub(r'[<>:"/\\|?*]', '_', solicitation_number)

            # Click Documents tab with timeout and error handling
            try:
                docs_tab = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#docs-itemsAbstractTab > a"))
                )
                docs_tab.click()
                self.random_delay(2, 3)
            except Exception as e:
                self.logger.info(f"No documents tab available: {str(e)}")
                return ""

            # Click all attachment links first - Fixed await issue
            click_script = """
                function clickAllAttachments() {
                    const links = [];
                    let rowIndex = 1;
                    
                    function sleep(ms) {
                        return new Promise(resolve => {
                            const start = Date.now();
                            while (Date.now() - start < ms) {}
                            resolve();
                        });
                    }
                    
                    while (true) {
                        try {
                            const xpath = `/html/body/main/div[1]/div[2]/div[3]/div[2]/div/div/table/tbody/tr[${rowIndex}]/td[1]/a`;
                            const result = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
                            const element = result.singleNodeValue;
                            
                            if (element) {
                                try {
                                    element.click();
                                    console.log(`Clicked attachment in row ${rowIndex}`);
                                    // Use synchronous delay instead of await
                                    sleep(500);
                                } catch (clickError) {
                                    console.log(`Click failed for row ${rowIndex}: ${clickError}`);
                                }
                                rowIndex++;
                            } else {
                                break;
                            }
                        } catch (e) {
                            console.log(`Error processing row ${rowIndex}: ${e}`);
                            break;
                        }
                    }
                    return rowIndex - 1;  // Return number of attachments clicked
                }
                return clickAllAttachments();
            """
            
            num_attachments = self.driver.execute_script(click_script)
            self.logger.info(f"Clicked {num_attachments} attachment links")
            self.random_delay(2, 3)  # Wait for downloads to start

            # Wait for downloads and move files
            downloaded_files = []
            start_time = time.time()
            timeout = 300  # 5 minutes total timeout

            while time.time() - start_time < timeout:
                # Get all files in temporary download folder
                temp_files = list(self.script_folder.glob('*'))
                
                for temp_file in temp_files:
                    if temp_file.is_file():
                        try:
                            # Wait for file size to stabilize
                            last_size = -1
                            current_size = temp_file.stat().st_size
                            
                            size_stable = False
                            size_check_start = time.time()
                            
                            while time.time() - size_check_start < 30:  # 30 seconds timeout for size check
                                if last_size == current_size:
                                    size_stable = True
                                    break
                                time.sleep(2)
                                last_size = current_size
                                if temp_file.exists():
                                    current_size = temp_file.stat().st_size
                                else:
                                    break

                            if size_stable and temp_file.exists():
                                # Create bid folder only when we have a stable file to move
                                bid_folder = self.script_folder_in_progress / clean_solicitation
                                bid_folder.mkdir(parents=True, exist_ok=True)
                                
                                # Use the actual filename from the downloaded file
                                actual_filename = temp_file.name
                                target_path = bid_folder / actual_filename
                                
                                if not target_path.exists():  # Only move if target doesn't exist
                                    safe_move(str(temp_file), str(target_path))
                                    downloaded_files.append(actual_filename)
                                    self.logger.info(f"✅ Moved file: {actual_filename}")
                                
                        except Exception as e:
                            self.logger.error(f"Error processing file {temp_file}: {str(e)}")
                            if temp_file.exists():
                                try:
                                    temp_file.unlink()
                                except:
                                    pass

                # Break if we've found all expected attachments
                if len(downloaded_files) >= num_attachments:
                    break
                
                time.sleep(1)

            # Clean up any remaining files in temp folder
            for temp_file in self.script_folder.glob('*'):
                try:
                    if temp_file.is_file():
                        temp_file.unlink()
                except:
                    pass

            if downloaded_files:
                # Update Excel with actual filenames
                self.update_excel_attachments(solicitation_number, ", ".join(downloaded_files))
                self.logger.info(f"Successfully downloaded attachments: {', '.join(downloaded_files)}")
            else:
                self.logger.warning("No attachments were successfully downloaded")
            
            return ", ".join(downloaded_files)

        except Exception as e:
            self.logger.error(f"Error in attachment download process: {str(e)}")
            return ""

    def update_excel_attachments(self, solicitation_number: str, attachments: str):
        """Update the Attachments column in Excel for the given solicitation number"""
        try:
            excel_path = self.script_folder_in_progress / f"{self.script_name}.xlsx"
            if not excel_path.exists():
                self.logger.error("Excel file not found")
                return

            df = pd.read_excel(excel_path)
            mask = df['Solicitation Number'] == solicitation_number
            if mask.any():
                df.loc[mask, 'Attachments'] = attachments
                df.to_excel(excel_path, index=False)
                self.logger.info(f"Updated Excel attachments for {solicitation_number}")
            else:
                self.logger.error(f"Solicitation number {solicitation_number} not found in Excel")

        except Exception as e:
            self.logger.error(f"Error updating Excel attachments: {str(e)}")

    def format_date(self, date_str: str) -> str:
        """Format date string to YYYY-MM-DD format"""
        try:
            if not date_str:
                return ""
            
            # Remove any time component
            date_part = date_str.split(" ")[0]
            
            # Check if already in YYYY-MM-DD format
            if re.match(r'^\d{4}-\d{2}-\d{2}$', date_part):
                self.logger.info("✓ Date is in correct format")
                return date_part
            
            # Try parsing as MM/DD/YYYY
            date_obj = datetime.strptime(date_part, "%m/%d/%Y")
            return date_obj.strftime("%Y-%m-%d")
                
        except Exception:
            return date_str

    def save_to_excel(self) -> bool:
        """Save bid data to Excel file with proper formatting"""
        try:
            # Save Excel in the _IN_PROGRESS folder
            excel_path = self.script_folder_in_progress / f"{self.script_name}.xlsx"

            # Convert bid data to DataFrame
            df = pd.DataFrame([asdict(bid) for bid in self.bids_data])

            # Define column mapping and order
            column_mapping = {
                "sl_no": "SL No",
                "posted_date": "Posted Date",
                "response_date": "Response Date",
                "notice_type": "Notice Type",
                "solicitation_number": "Solicitation Number",
                "solicitation_title": "Solicitation Title",
                "agency": "Agency",
                "category": "Category",
                "description": "Description",
                "additional_summary": "Additional Summary",
                "contracting_office_address": "Contracting Office Address",
                "contact_information": "Contact Information",
                "bid_detail_page_url": "Bid Detail Page URL",
                "attachments": "Attachments"
            }

            # Rename columns
            df = df.rename(columns=column_mapping)

            # Reorder columns to match required order
            ordered_columns = [
                "SL No",
                "Posted Date",
                "Response Date",
                "Notice Type",
                "Solicitation Number",
                "Solicitation Title", 
                "Agency",
                "Category",
                "Description",
                "Additional Summary",
                "Contracting Office Address",
                "Contact Information",
                "Bid Detail Page URL",
                "Attachments"
            ]
            df = df[ordered_columns]

            # Write to Excel
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Bids")

                # Auto-adjust column widths
                worksheet = writer.sheets["Bids"]
                for idx, col in enumerate(df.columns):
                    max_length = max(df[col].astype(str).apply(len).max(), len(str(col))) + 2
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 100)

                # Set text format for all cells
                for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.number_format = "@"

            self.logger.info(f"✅ Successfully saved data to Excel: {excel_path}")
            return True

        except Exception as e:
            self.logger.error(f"❌ Error saving to Excel: {str(e)}")
            return False



    def update_excel_after_bid(self, bid_data: BidData) -> bool:
        """Update Excel file after each bid is processed"""
        try:
            self.bids_data.append(bid_data)
            if self.save_to_excel():
                self.logger.info(
                    f"✅ Successfully updated Excel with bid: {bid_data.solicitation_number}"
                )
                return True
            return False
        except Exception as e:
            self.logger.error(f"❌ Error updating Excel: {str(e)}")
            return False


def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(description="BidNet Direct Scraper")
    parser.add_argument(
        "--days",
        type=int,
        default=2,
        help="Number of days to look back for bids (default: 2)",
    )
    args = parser.parse_args()

    print("🟢 BidNet Direct Scraping Started")

    try:
        with BidNetScraper(days=args.days) as scraper:
            # Login to the site
            if not scraper.login():
                print("❌ Login failed, exiting...")
                return

            # Navigate to purchasing groups
            if not scraper.navigate_to_purchasing_groups():
                print("❌ Failed to navigate to purchasing groups")
                return

            # Get all purchasing groups
            groups = scraper.get_purchasing_groups()
            print(f"📋 Found {len(groups)} purchasing groups")

            # Process each group
            for group in groups:
                print(f"\n🔍 Processing group: {group['name']}")

                # Navigate to group page
                scraper.driver.get(group["url"])
                scraper.random_delay(2, 4)

                # Check if group bids tab exists
                bid_count = scraper.check_group_bids_tab()
                if bid_count is None:
                    print(f"⏭️ Skipping {group['name']}, no Group Bids tab found")
                    continue

                print(f"📊 Found {bid_count} bids in {group['name']}")

                # Extract and process bid links
                bid_links = scraper.extract_bid_links()
                if bid_links:
                    print(f"🔗 Found {len(bid_links)} bid links to process")
                    scraper.process_bid_links(bid_links)
                else:
                    print("⚠️ No bid links found in date range")

            print("\n🎉 All Bids and Attachments Extraction Successfully Completed")

    except Exception as e:
        print(f"❌ Critical error: {str(e)}")
        play_notification_sound()
        input("Press Enter to continue...")


if __name__ == "__main__":
    main()
